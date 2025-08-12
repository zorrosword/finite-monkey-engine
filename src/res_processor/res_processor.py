import pandas as pd
from tqdm import tqdm
import json
from openai_api.openai import ask_claude, ask_deepseek, ask_o3_mini_json, common_ask_for_json,ask_claude_37
import concurrent.futures
from threading import Lock
import math
import re
import time
import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

class ResProcessor:
    def __init__(self, df, max_group_size=10, iteration_rounds=2, enable_chinese_translation=False):
        """
        åˆå§‹åŒ–ResProcessor
        
        Args:
            df: åŒ…å«æ¼æ´æ•°æ®çš„DataFrame
            max_group_size: æ¯ç»„æœ€å¤§æ¼æ´æ•°é‡ï¼Œé»˜è®¤ä¸º10
            iteration_rounds: è¿­ä»£è½®æ•°ï¼Œé»˜è®¤ä¸º2
            enable_chinese_translation: æ˜¯å¦å¯ç”¨ä¸­æ–‡ç¿»è¯‘ï¼Œé»˜è®¤ä¸ºFalse
        """
        self.df = df
        self.lock = Lock()
        self.max_group_size = max_group_size
        self.iteration_rounds = iteration_rounds
        self.enable_chinese_translation = enable_chinese_translation
        
        print(f"ResProcessoråˆå§‹åŒ–:")
        print(f"  - æœ€å¤§ç»„å¤§å°: {self.max_group_size}")
        print(f"  - è¿­ä»£è½®æ•°: {self.iteration_rounds}")
        print(f"  - ä¸­æ–‡ç¿»è¯‘: {'å¯ç”¨' if self.enable_chinese_translation else 'ç¦ç”¨'}")

    def process(self):
        """ä¸»å¤„ç†å‡½æ•°ï¼Œå®ç°å¤šè½®è¿­ä»£çš„æ¼æ´å½’é›†"""
        print("å¼€å§‹æ¼æ´å½’é›†å¤„ç†...")
        print(f"æ€»æ¼æ´æ•°é‡: {len(self.df)}")
        
        # æ·»åŠ è¾…åŠ©åˆ—
        self.df['flow_code_len'] = self.df['ä¸šåŠ¡æµç¨‹ä»£ç '].str.len()
        
        # ç¬¬ä¸€æ­¥ï¼šæŒ‰ä¸šåŠ¡æµç¨‹ä»£ç åˆ†ç»„
        initial_groups = list(self.df.groupby('ä¸šåŠ¡æµç¨‹ä»£ç '))
        print(f"åˆå§‹åˆ†ç»„æ•°é‡: {len(initial_groups)}")
        
        # æ‰“å°åˆå§‹åˆ†ç»„è¯¦æƒ…
        print("\n=== åˆå§‹åˆ†ç»„è¯¦æƒ… ===")
        for i, (flow_code, group) in enumerate(initial_groups):
            flow_code_preview = flow_code[:100] + "..." if len(flow_code) > 100 else flow_code
            print(f"åˆ†ç»„ {i+1}: ä¸šåŠ¡æµç¨‹ä»£ç é•¿åº¦={len(flow_code)}, æ¼æ´æ•°é‡={len(group)}")
            print(f"  ä»£ç é¢„è§ˆ: {flow_code_preview}")
            if len(group) > self.max_group_size:
                print(f"  âš ï¸  è¯¥åˆ†ç»„è¶…è¿‡æœ€å¤§é™åˆ¶({self.max_group_size})ï¼Œéœ€è¦ç»†åˆ†")
        
        # ç¬¬äºŒæ­¥ï¼šç»†åˆ†å¤§ç»„ï¼ˆè¶…è¿‡10ä¸ªæ¼æ´çš„ç»„ï¼‰
        refined_groups = self._refine_large_groups(initial_groups)
        print(f"\nç»†åˆ†ååˆ†ç»„æ•°é‡: {len(refined_groups)}")
        
        # ç¬¬ä¸‰æ­¥ï¼šå¤šè½®è¿­ä»£å½’é›†
        current_groups = refined_groups
        for round_num in range(self.iteration_rounds):
            print(f"\n{'='*50}")
            print(f"å¼€å§‹ç¬¬ {round_num + 1} è½®å½’é›†")
            print(f"{'='*50}")
            print(f"è¾“å…¥åˆ†ç»„æ•°é‡: {len(current_groups)}")
            
            # æ‰“å°å½“å‰è½®æ¬¡åˆ†ç»„è¯¦æƒ…
            total_vulns = sum(len(group) for group in current_groups)
            print(f"å½“å‰è½®æ¬¡æ€»æ¼æ´æ•°é‡: {total_vulns}")
            for i, group in enumerate(current_groups):
                print(f"  åˆ†ç»„ {i+1}: {len(group)} ä¸ªæ¼æ´")
            
            current_groups = self._iteration_round(current_groups, round_num + 1)
            
            # æ£€æŸ¥è¿”å›çš„æ•°æ®ç±»å‹å¹¶å¤„ç†æå‰åœæ­¢
            if isinstance(current_groups, list) and len(current_groups) > 0:
                if isinstance(current_groups[0], dict):
                    # è¿”å›çš„æ˜¯å­—å…¸åˆ—è¡¨ï¼Œè¯´æ˜æ˜¯æœ€ç»ˆç»“æœ
                    print(f"ç¬¬ {round_num + 1} è½®å½’é›†å®Œæˆï¼Œè¿”å›æœ€ç»ˆç»“æœ: {len(current_groups)} ä¸ªæ¼æ´æŠ¥å‘Š")
                    break
                else:
                    # è¿”å›çš„æ˜¯DataFrameåˆ—è¡¨ï¼Œç»§ç»­ä¸‹ä¸€è½®
                    print(f"ç¬¬ {round_num + 1} è½®å½’é›†ååˆ†ç»„æ•°é‡: {len(current_groups)}")
            else:
                print(f"ç¬¬ {round_num + 1} è½®å½’é›†å®Œæˆï¼Œæ— æœ‰æ•ˆç»“æœ")
                break
        
        # ç¬¬å››æ­¥ï¼šæ„å»ºæœ€ç»ˆç»“æœ
        final_results = self._build_final_results(current_groups)
        print(f"\næœ€ç»ˆç»“æœ: {len(final_results)} ä¸ªæ¼æ´æŠ¥å‘Š")
        
        # ç¬¬äº”æ­¥ï¼šä¸­æ–‡ç¿»è¯‘ï¼ˆå¯é€‰ï¼‰
        final_results = self._translate_to_chinese(final_results)
        
        # æ¸…ç†è¾…åŠ©åˆ—å¹¶è¿”å›ç»“æœ
        new_df = pd.DataFrame(final_results)
        if 'flow_code_len' in new_df.columns:
            new_df = new_df.drop('flow_code_len', axis=1)
            
        original_columns = [col for col in self.df.columns if col != 'flow_code_len']
        new_df = new_df[original_columns]
        
        return new_df

    def _refine_large_groups(self, initial_groups):
        """ç»†åˆ†å¤§ç»„ï¼Œç¡®ä¿æ¯ç»„ä¸è¶…è¿‡æœ€å¤§é™åˆ¶"""
        refined_groups = []
        
        print("\n=== å¼€å§‹ç»†åˆ†å¤§ç»„ ===")
        
        # ä½¿ç”¨å¤šçº¿ç¨‹å¤„ç†å¤§ç»„ç»†åˆ†
        with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
            future_to_group = {
                executor.submit(self._process_single_group, i, flow_code, group): (i, flow_code, group)
                for i, (flow_code, group) in enumerate(initial_groups)
            }
            
            with tqdm(total=len(initial_groups), desc="ç»†åˆ†å¤§ç»„") as pbar:
                for future in concurrent.futures.as_completed(future_to_group):
                    try:
                        subgroups = future.result()
                        refined_groups.extend(subgroups)
                    except Exception as e:
                        i, flow_code, group = future_to_group[future]
                        print(f"åˆ†ç»„ {i+1} ç»†åˆ†å¤±è´¥: {str(e)}")
                        refined_groups.append(group)
                    pbar.update(1)
        
        print(f"ç»†åˆ†å®Œæˆ: {len(initial_groups)} ä¸ªåˆå§‹åˆ†ç»„ -> {len(refined_groups)} ä¸ªç»†åˆ†ååˆ†ç»„")
        return refined_groups

    def _process_single_group(self, index, flow_code, group):
        """å¤„ç†å•ä¸ªåˆ†ç»„çš„ç»†åˆ†"""
        if len(group) <= self.max_group_size:
            flow_code_preview = flow_code[:50] + "..." if len(flow_code) > 50 else flow_code
            print(f"åˆ†ç»„ {index+1} (ä¸šåŠ¡æµç¨‹ä»£ç : {flow_code_preview}): å¤§å° {len(group)} <= {self.max_group_size}ï¼Œæ— éœ€ç»†åˆ†")
            return [group]
        else:
            # å°†å¤§ç»„æ‹†åˆ†ä¸ºå°ç»„
            num_subgroups = math.ceil(len(group) / self.max_group_size)
            flow_code_preview = flow_code[:50] + "..." if len(flow_code) > 50 else flow_code
            print(f"åˆ†ç»„ {index+1} (ä¸šåŠ¡æµç¨‹ä»£ç : {flow_code_preview}): å¤§å° {len(group)} > {self.max_group_size}ï¼Œéœ€è¦æ‹†åˆ†ä¸º {num_subgroups} ä¸ªå­ç»„")
            
            group_list = group.to_dict('records')
            subgroups = []
            
            for j in range(num_subgroups):
                start_idx = j * self.max_group_size
                end_idx = min(start_idx + self.max_group_size, len(group_list))
                subgroup_data = group_list[start_idx:end_idx]
                subgroup_df = pd.DataFrame(subgroup_data)
                subgroups.append(subgroup_df)
                print(f"  å­ç»„ {j+1}: æ¼æ´æ•°é‡ {len(subgroup_data)} (è¡Œ {start_idx+1}-{end_idx})")
            
            return subgroups

    def _iteration_round(self, groups, round_num):
        """æ‰§è¡Œä¸€è½®è¿­ä»£å½’é›†"""
        print(f"\n--- ç¬¬ {round_num} è½®å¤„ç†å¼€å§‹ ---")
        print(f"å¤„ç† {len(groups)} ä¸ªåˆ†ç»„")
        
        # æ‰“å°æ¯ä¸ªåˆ†ç»„çš„è¯¦æƒ…
        for i, group in enumerate(groups):
            print(f"  è¾“å…¥åˆ†ç»„ {i+1}: {len(group)} ä¸ªæ¼æ´")
        
        # æ­¥éª¤1ï¼šå¯¹æ¯ä¸ªç»„è¿›è¡Œæ¼æ´åˆ†ç±»
        print(f"\næ­¥éª¤1: å¼€å§‹æ¼æ´åˆ†ç±»...")
        classified_groups = []
        with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
            future_to_group = {
                executor.submit(self._classify_vulnerabilities_in_group, group): (group, i) 
                for i, group in enumerate(groups)
            }
            
            with tqdm(total=len(groups), desc=f"ç¬¬{round_num}è½®-æ¼æ´åˆ†ç±»") as pbar:
                for future in concurrent.futures.as_completed(future_to_group):
                    try:
                        classified_result = future.result()
                        original_group, group_index = future_to_group[future]
                        
                        print(f"  åˆ†ç»„ {group_index+1} åˆ†ç±»ç»“æœ: {len(original_group)} -> {len(classified_result)} ä¸ªå­ç»„")
                        for j, subgroup in enumerate(classified_result):
                            print(f"    å­ç»„ {j+1}: {len(subgroup)} ä¸ªæ¼æ´")
                        
                        classified_groups.extend(classified_result)
                    except Exception as e:
                        original_group, group_index = future_to_group[future]
                        print(f"  åˆ†ç»„ {group_index+1} åˆ†ç±»å¤±è´¥ï¼Œä¿æŒåŸç»„: {str(e)}")
                        classified_groups.append(original_group)
                    pbar.update(1)
        
        print(f"åˆ†ç±»å®Œæˆ: {len(groups)} ä¸ªè¾“å…¥åˆ†ç»„ -> {len(classified_groups)} ä¸ªåˆ†ç±»ååˆ†ç»„")
        
        # æ­¥éª¤2ï¼šå¯¹åˆ†ç±»åçš„ç»„è¿›è¡Œå»é‡å¤„ç†ï¼ˆæ¯ç»„åªä¿ç•™ç¬¬ä¸€ä¸ªæ¼æ´ï¼‰
        print(f"\næ­¥éª¤2: å¼€å§‹å»é‡å¤„ç†...")
        deduplicated_results = []
        
        for i, group in enumerate(classified_groups):
            # è¾“å…¥éªŒè¯ï¼šç¡®ä¿groupæ˜¯DataFrame
            if not isinstance(group, pd.DataFrame):
                print(f"  åˆ†ç»„ {i+1}: âš ï¸  è¾“å…¥ä¸æ˜¯DataFrameï¼Œç±»å‹: {type(group)}")
                if isinstance(group, dict):
                    # å¦‚æœæ˜¯å•ä¸ªå­—å…¸ï¼Œç›´æ¥ä¿ç•™
                    print(f"  åˆ†ç»„ {i+1}: å•ä¸ªå­—å…¸ï¼Œç›´æ¥ä¿ç•™")
                    deduplicated_results.append(group)
                    continue
                elif isinstance(group, list):
                    # å¦‚æœæ˜¯å­—å…¸åˆ—è¡¨ï¼Œè½¬æ¢ä¸ºDataFrame
                    group = pd.DataFrame(group)
                else:
                    print(f"  åˆ†ç»„ {i+1}: âŒ æ— æ³•å¤„ç†çš„è¾“å…¥ç±»å‹ï¼Œè·³è¿‡")
                    continue
            
            if len(group) <= 1:
                # å•ä¸ªæ¼æ´ï¼Œç›´æ¥ä¿ç•™
                print(f"  åˆ†ç»„ {i+1}: å•ä¸ªæ¼æ´ï¼Œç›´æ¥ä¿ç•™")
                preserved_vuln = group.iloc[0].to_dict()
                deduplicated_results.append(preserved_vuln)
            else:
                # å¤šä¸ªæ¼æ´ï¼Œåªä¿ç•™ç¬¬ä¸€ä¸ª
                first_vuln = group.iloc[0]
                removed_count = len(group) - 1
                print(f"  åˆ†ç»„ {i+1}: {len(group)} ä¸ªæ¼æ´ -> ä¿ç•™ç¬¬1ä¸ªï¼Œåˆ é™¤ {removed_count} ä¸ªé‡å¤é¡¹")
                
                # è®°å½•è¢«åˆ é™¤çš„æ¼æ´ID
                removed_ids = [str(row['ID']) for _, row in group.iloc[1:].iterrows()]
                print(f"    åˆ é™¤çš„æ¼æ´ID: {', '.join(removed_ids)}")
                print(f"    ä¿ç•™çš„æ¼æ´ID: {first_vuln['ID']}")
                
                preserved_vuln = first_vuln.to_dict()
                deduplicated_results.append(preserved_vuln)
        
        print(f"å»é‡å®Œæˆ: {len(classified_groups)} ä¸ªåˆ†ç±»ååˆ†ç»„ -> {len(deduplicated_results)} ä¸ªå»é‡åç»“æœ")
        
        # ç»Ÿè®¡å»é‡æ•ˆæœ
        original_total = sum(len(group) for group in classified_groups)
        final_total = len(deduplicated_results)
        removed_total = original_total - final_total
        print(f"å»é‡ç»Ÿè®¡: åŸå§‹ {original_total} ä¸ªæ¼æ´ -> æœ€ç»ˆ {final_total} ä¸ªæ¼æ´ï¼Œåˆ é™¤äº† {removed_total} ä¸ªé‡å¤é¡¹")
        
        # æ£€æŸ¥æ˜¯å¦éœ€è¦æå‰åœæ­¢
        if final_total < 40:
            print(f"\nğŸ¯ å½’é›†ç»“æœå·²å°‘äº25ä¸ª({final_total}ä¸ª)ï¼Œæå‰åœæ­¢è¿­ä»£")
            print(f"ç¬¬ {round_num} è½®ä¸ºæå‰ç»“æŸè½®æ¬¡")
            return deduplicated_results  # è¿”å›å­—å…¸åˆ—è¡¨ä½œä¸ºæœ€ç»ˆç»“æœ
        
        # æ­¥éª¤3ï¼šå‡†å¤‡ä¸‹ä¸€è½®çš„åˆ†ç»„ï¼ˆæ¯10ä¸ªä¸€ç»„ï¼‰
        if round_num < self.iteration_rounds:
            print(f"\næ­¥éª¤3: å‡†å¤‡ç¬¬ {round_num+1} è½®åˆ†ç»„...")
            next_round_groups = self._prepare_next_round_groups_simple(deduplicated_results)
            print(f"ä¸ºä¸‹ä¸€è½®å‡†å¤‡äº† {len(next_round_groups)} ä¸ªåˆ†ç»„")
            
            # æ‰“å°ä¸‹ä¸€è½®åˆ†ç»„è¯¦æƒ…
            total_items = sum(len(group) for group in next_round_groups)
            print(f"ä¸‹ä¸€è½®æ€»é¡¹ç›®æ•°: {total_items}")
            for i, group in enumerate(next_round_groups):
                print(f"  ä¸‹ä¸€è½®åˆ†ç»„ {i+1}: {len(group)} ä¸ªé¡¹ç›®")
            
            return next_round_groups  # è¿”å›DataFrameåˆ—è¡¨ç»§ç»­ä¸‹ä¸€è½®
        else:
            print(f"\nç¬¬ {round_num} è½®ä¸ºæœ€åä¸€è½®ï¼Œä¸éœ€è¦å‡†å¤‡ä¸‹ä¸€è½®åˆ†ç»„")
            return deduplicated_results  # è¿”å›å­—å…¸åˆ—è¡¨ä½œä¸ºæœ€ç»ˆç»“æœ

    def _classify_vulnerabilities_in_group(self, group):
        """ä½¿ç”¨æ–°çš„promptå¯¹å•ä¸ªç»„è¿›è¡Œæ¼æ´åˆ†ç±»"""
        # è¾“å…¥éªŒè¯ï¼šç¡®ä¿groupæ˜¯DataFrame
        if not isinstance(group, pd.DataFrame):
            print(f"    âš ï¸  è¾“å…¥ä¸æ˜¯DataFrameï¼Œç±»å‹: {type(group)}")
            if isinstance(group, dict):
                # å¦‚æœæ˜¯å•ä¸ªå­—å…¸ï¼Œè½¬æ¢ä¸ºDataFrame
                group = pd.DataFrame([group])
            elif isinstance(group, list):
                # å¦‚æœæ˜¯å­—å…¸åˆ—è¡¨ï¼Œè½¬æ¢ä¸ºDataFrame
                group = pd.DataFrame(group)
            else:
                print(f"    âŒ æ— æ³•å¤„ç†çš„è¾“å…¥ç±»å‹ï¼Œç›´æ¥è¿”å›")
                return [group]
        
        if len(group) <= 1:
            print(f"    å•ä¸ªæ¼æ´æ— éœ€åˆ†ç±»ï¼Œç›´æ¥è¿”å›")
            return [group]
        
        print(f"    å¼€å§‹åˆ†ç±» {len(group)} ä¸ªæ¼æ´...")
        
        # æ„å»ºæ¼æ´ä¿¡æ¯
        vuln_keys = []
        vuln_descriptions = []
        
        for _, row in group.iterrows():
            key = str(row['UUID'])
            description = f"æ¼æ´å†…å®¹ï¼š{row['æ¼æ´ç»“æœ']}"
            vuln_keys.append(key)
            vuln_descriptions.append(f"Key: {key}\n{description}")
        
        print(f"    æ„å»ºäº† {len(vuln_keys)} ä¸ªæ¼æ´çš„æè¿°ä¿¡æ¯")
        
        classification_prompt = f"""# Role: ç½‘ç»œå®‰å…¨æ¼æ´åˆ†ç±»ä¸“å®¶

## Profile
- language: ä¸­æ–‡
- description: ä¸“ä¸šçš„ç½‘ç»œå®‰å…¨æ¼æ´è¯†åˆ«ä¸åˆ†ç±»ä¸“å®¶ï¼Œå…·å¤‡æ·±åšçš„å®‰å…¨æŠ€æœ¯èƒŒæ™¯å’Œä¸°å¯Œçš„æ¼æ´åˆ†æç»éªŒï¼Œèƒ½å¤Ÿå‡†ç¡®è¯†åˆ«å„ç±»å®‰å…¨æ¼æ´çš„æœ¬è´¨ç‰¹å¾ï¼Œå¹¶è¿›è¡Œç²¾å‡†åˆ†ç±»
- background: æ‹¥æœ‰å¤šå¹´ç½‘ç»œå®‰å…¨ä»ä¸šç»éªŒï¼Œç²¾é€šå„ç§æ¼æ´ç±»å‹ã€æ”»å‡»åŸç†å’Œé˜²æŠ¤æœºåˆ¶ï¼Œå‚ä¸è¿‡å¤§é‡æ¼æ´æŒ–æ˜ã€åˆ†æå’Œä¿®å¤å·¥ä½œ
- personality: ä¸¥è°¨ç»†è‡´ã€é€»è¾‘æ¸…æ™°ã€ä¸“ä¸šä¸“æ³¨ã€è¿½æ±‚ç²¾ç¡®

## Skills

1. æ¼æ´è¯†åˆ«ä¸åˆ†æ
   - æ¼æ´æœ¬è´¨åˆ¤æ–­: èƒ½å¤Ÿé€è¿‡è¡¨é¢ç°è±¡è¯†åˆ«æ¼æ´çš„æ ¹æœ¬åŸç†å’Œæˆå› 
   - æ¼æ´ç±»å‹å½’çº³: ç²¾é€šOWASPã€CVEã€CWEç­‰æ ‡å‡†åˆ†ç±»ä½“ç³»
   - æ”»å‡»å‘é‡åˆ†æ: æ·±åº¦ç†è§£å„ç§æ”»å‡»æ‰‹æ®µå’Œåˆ©ç”¨æ–¹å¼
   - å½±å“èŒƒå›´è¯„ä¼°: å‡†ç¡®åˆ¤æ–­æ¼æ´çš„å±å®³ç¨‹åº¦å’Œå½±å“èŒƒå›´

2. åˆ†ç±»ä¸æ ‡å‡†åŒ–
   - æ ‡å‡†åŒ–åˆ†ç±»: åŸºäºå›½é™…æ ‡å‡†å’Œè¡Œä¸šæœ€ä½³å®è·µè¿›è¡Œåˆ†ç±»
   - ç›¸ä¼¼æ€§è¯†åˆ«: å¿«é€Ÿè¯†åˆ«ä¸åŒè¡¨è¿°ä¸‹çš„ç›¸åŒæœ¬è´¨æ¼æ´
   - å±‚æ¬¡åŒ–å½’ç±»: å»ºç«‹æ¸…æ™°çš„åˆ†ç±»å±‚æ¬¡å’Œé€»è¾‘å…³ç³»
   - ç»“æœæ ¼å¼åŒ–: æŒ‰ç…§æŒ‡å®šæ ¼å¼å‡†ç¡®è¾“å‡ºåˆ†ç±»ç»“æœ

## Rules

1. åˆ†ç±»åŸºæœ¬åŸåˆ™:
   - æœ¬è´¨ç›¸åŒåŸåˆ™: åªæœ‰æ¼æ´çš„æ ¹æœ¬åŸç†å’Œæˆå› å®Œå…¨ç›¸åŒæ‰å½’ä¸ºä¸€ç±»
   - ä¸¥æ ¼åŒºåˆ†åŸåˆ™: è¡¨é¢ç›¸ä¼¼ä½†æœ¬è´¨ä¸åŒçš„æ¼æ´å¿…é¡»åˆ†ä¸ºä¸åŒç±»åˆ«
   - å®Œæ•´æ€§åŸåˆ™: ç¡®ä¿æ‰€æœ‰æä¾›çš„æ¼æ´keyéƒ½è¢«æ­£ç¡®åˆ†ç±»ï¼Œä¸é—æ¼ä¸é‡å¤
   - ä¸€è‡´æ€§åŸåˆ™: é‡‡ç”¨ç»Ÿä¸€çš„åˆ†ç±»æ ‡å‡†ï¼Œç¡®ä¿åˆ†ç±»ç»“æœçš„ä¸€è‡´æ€§

2. åˆ†æè¡Œä¸ºå‡†åˆ™:
   - æ·±åº¦åˆ†æ: æ·±å…¥åˆ†ææ¯ä¸ªæ¼æ´çš„æŠ€æœ¯ç»†èŠ‚å’Œæ”»å‡»åŸç†
   - å¯¹æ¯”éªŒè¯: é€šè¿‡å¤šç»´åº¦å¯¹æ¯”ç¡®è®¤æ¼æ´çš„ç›¸ä¼¼æ€§å’Œå·®å¼‚æ€§
   - å®¢è§‚åˆ¤æ–­: åŸºäºæŠ€æœ¯äº‹å®è¿›è¡Œåˆ†ç±»ï¼Œé¿å…ä¸»è§‚è‡†æ–­
   - ç²¾å‡†è¯†åˆ«: å‡†ç¡®è¯†åˆ«æ¼æ´çš„æ ¸å¿ƒç‰¹å¾å’Œå…³é”®å·®å¼‚ç‚¹

3. è¾“å‡ºé™åˆ¶æ¡ä»¶:
   - æ ¼å¼ä¸¥æ ¼: ä¸¥æ ¼æŒ‰ç…§è§„å®šçš„JSONæ ¼å¼è¾“å‡ºåˆ†ç±»ç»“æœ
   - å†…å®¹çº¯å‡€: ä»…è¾“å‡ºåˆ†ç±»ç»“æœï¼Œä¸åŒ…å«ä»»ä½•è§£é‡Šè¯´æ˜æˆ–é¢å¤–ä¿¡æ¯
   - æ ‡è¯†å”¯ä¸€: ç¡®ä¿æ¯ä¸ªåˆ†ç±»ç»„æœ‰å”¯ä¸€çš„åˆ†ç»„æ ‡è¯†
   - æ— å†—ä½™ä¿¡æ¯: ä¸è¾“å‡ºåˆ†æè¿‡ç¨‹ã€ç†ç”±è¯´æ˜æˆ–å…¶ä»–è¾…åŠ©å†…å®¹

## Workflows

- ç›®æ ‡: å°†è¾“å…¥çš„æ¼æ´æŒ‰ç…§æœ¬è´¨ç‰¹å¾è¿›è¡Œç²¾å‡†åˆ†ç±»
- æ­¥éª¤ 1: æ¥æ”¶å¹¶è§£ææ‰€æœ‰æ¼æ´keyï¼Œç†è§£æ¯ä¸ªæ¼æ´çš„æŠ€æœ¯ç‰¹å¾å’Œæ”»å‡»åŸç†
- æ­¥éª¤ 2: åŸºäºæ¼æ´çš„æ ¹æœ¬æˆå› ã€æ”»å‡»æœºåˆ¶ã€åˆ©ç”¨æ–¹å¼ç­‰æ ¸å¿ƒè¦ç´ è¿›è¡Œæ·±åº¦åˆ†æ
- æ­¥éª¤ 3: è¯†åˆ«å…·æœ‰ç›¸åŒæœ¬è´¨ç‰¹å¾çš„æ¼æ´ï¼Œå°†å…¶å½’ä¸ºåŒä¸€ç±»åˆ«ï¼Œä¸ºæ¯ä¸ªç±»åˆ«åˆ†é…å”¯ä¸€æ ‡è¯†
- æ­¥éª¤ 4: æŒ‰ç…§æ ‡å‡†JSONæ ¼å¼è¾“å‡ºæœ€ç»ˆåˆ†ç±»ç»“æœ
- é¢„æœŸç»“æœ: è¾“å‡ºæ ‡å‡†åŒ–çš„æ¼æ´åˆ†ç±»ç»“æœï¼Œæ ¼å¼ä¸º{{"group_1":["key1","key2"],"group_2":["key3","key4"],"group_3":["key5"]}}

## Initialization
ä½œä¸ºç½‘ç»œå®‰å…¨æ¼æ´åˆ†ç±»ä¸“å®¶ï¼Œä½ å¿…é¡»éµå®ˆä¸Šè¿°Rulesï¼ŒæŒ‰ç…§Workflowsæ‰§è¡Œä»»åŠ¡ã€‚

ä»¥ä¸‹æ˜¯éœ€è¦åˆ†ç±»çš„æ¼æ´ä¿¡æ¯ï¼š
{chr(10).join(vuln_descriptions)}"""

        try:
            # è°ƒç”¨åˆ†ç±»API
            print(f"    è°ƒç”¨AIè¿›è¡Œæ¼æ´åˆ†ç±»...")
            classification_result = ask_claude(classification_prompt)
            classification_data = json.loads(self._extract_json_from_text(classification_result))
            
            print(f"    AIåˆ†ç±»ç»“æœ: {len(classification_data)} ä¸ªåˆ†ç»„")
            for group_name, keys in classification_data.items():
                print(f"      {group_name}: {len(keys)} ä¸ªæ¼æ´")
            
            # æ ¹æ®åˆ†ç±»ç»“æœæ„å»ºæ–°çš„åˆ†ç»„
            new_groups = []
            for group_name, keys in classification_data.items():
                # æ ¹æ®keysç­›é€‰å¯¹åº”çš„è¡Œ
                group_rows = group[group['UUID'].astype(str).isin([str(k) for k in keys])]
                if not group_rows.empty:
                    new_groups.append(group_rows)
                    print(f"      åˆ›å»ºåˆ†ç»„ {group_name}: {len(group_rows)} ä¸ªæ¼æ´")
            
            if not new_groups:
                print(f"    åˆ†ç±»å¤±è´¥ï¼Œè¿”å›åŸåˆ†ç»„")
                return [group]
            
            print(f"    åˆ†ç±»æˆåŠŸ: {len(group)} -> {len(new_groups)} ä¸ªå­ç»„")
            return new_groups
            
        except Exception as e:
            print(f"    æ¼æ´åˆ†ç±»å¤±è´¥: {str(e)}")
            return [group]

    def _prepare_next_round_groups_simple(self, results_list):
        """ç®€åŒ–ç‰ˆçš„ä¸‹ä¸€è½®åˆ†ç»„å‡†å¤‡ï¼ˆå¤„ç†ç»“æœå­—å…¸åˆ—è¡¨ï¼‰"""
        print(f"    ä¸ºä¸‹ä¸€è½®é‡æ–°åˆ†ç»„ {len(results_list)} ä¸ªç»“æœ...")
        
        # ç›´æ¥æŒ‰é¡ºåºè¿›è¡Œåˆ†ç»„ï¼Œä¸å†æ’åº
        print(f"    æŒ‰é¡ºåºè¿›è¡Œåˆ†ç»„...")
        
        # å°†ç»“æœåˆ—è¡¨æŒ‰æœ€å¤§ç»„å¤§å°è¿›è¡Œåˆ†ç»„
        next_round_groups = []
        
        # ä½¿ç”¨å¤šçº¿ç¨‹åˆ›å»ºåˆ†ç»„
        with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
            group_futures = []
            
            for i in range(0, len(results_list), self.max_group_size):
                group_data = results_list[i:i + self.max_group_size]
                future = executor.submit(self._create_group, i, group_data, len(results_list))
                group_futures.append(future)
            
            with tqdm(total=len(group_futures), desc="åˆ›å»ºä¸‹ä¸€è½®åˆ†ç»„") as pbar:
                for future in concurrent.futures.as_completed(group_futures):
                    try:
                        group_df = future.result()
                        next_round_groups.append(group_df)
                    except Exception as e:
                        print(f"åˆ›å»ºåˆ†ç»„å¤±è´¥: {str(e)}")
                    pbar.update(1)
        
        # æŒ‰ç´¢å¼•æ’åºç¡®ä¿é¡ºåºæ­£ç¡®
        next_round_groups.sort(key=lambda x: x.index[0] if len(x) > 0 else 0)
        
        print(f"    ä¸ºä¸‹ä¸€è½®åˆ›å»ºäº† {len(next_round_groups)} ä¸ªåˆ†ç»„")
        return next_round_groups

    def _create_group(self, start_index, group_data, total_len):
        """åˆ›å»ºå•ä¸ªåˆ†ç»„"""
        group_df = pd.DataFrame(group_data)
        
        start_idx = start_index + 1
        end_idx = min(start_index + self.max_group_size, total_len)
        
        print(f"      åˆ›å»ºä¸‹ä¸€è½®åˆ†ç»„: é¡¹ç›® {start_idx}-{end_idx} ({len(group_data)} ä¸ª)")
        
        return group_df

    def _build_final_results(self, final_groups):
        """æ„å»ºæœ€ç»ˆç»“æœï¼ˆç®€åŒ–ç‰ˆï¼Œå¤„ç†ç»“æœå­—å…¸åˆ—è¡¨ï¼‰"""
        print(f"\n=== æ„å»ºæœ€ç»ˆç»“æœ ===")
        
        # å¦‚æœfinal_groupsæ˜¯å­—å…¸åˆ—è¡¨ï¼Œç›´æ¥è¿”å›
        if isinstance(final_groups, list) and all(isinstance(item, dict) for item in final_groups):
            print(f"æœ€ç»ˆç»“æœå·²æ˜¯å­—å…¸åˆ—è¡¨æ ¼å¼: {len(final_groups)} ä¸ªæ¼æ´æŠ¥å‘Š")
            print(f"  - æ‰€æœ‰æŠ¥å‘Šå‡ä¸ºç‹¬ç«‹æŠ¥å‘Šï¼ˆå·²å»é‡ï¼‰")
            return final_groups
        
        # å¦åˆ™æŒ‰åŸé€»è¾‘å¤„ç†
        print(f"å¤„ç† {len(final_groups)} ä¸ªæœ€ç»ˆåˆ†ç»„...")
        all_results = []
        
        for i, group_result in enumerate(final_groups):
            if isinstance(group_result, dict):
                all_results.append(group_result)
                print(f"  æœ€ç»ˆåˆ†ç»„ {i+1}: å­—å…¸ç±»å‹ (1 ä¸ªç‹¬ç«‹æŠ¥å‘Š)")
            elif isinstance(group_result, pd.DataFrame):
                records = group_result.to_dict('records')
                all_results.extend(records)
                print(f"  æœ€ç»ˆåˆ†ç»„ {i+1}: DataFrameç±»å‹ ({len(records)} ä¸ªç‹¬ç«‹æŠ¥å‘Š)")
            else:
                if hasattr(group_result, 'to_dict'):
                    all_results.append(group_result.to_dict())
                    print(f"  æœ€ç»ˆåˆ†ç»„ {i+1}: å…¶ä»–ç±»å‹è½¬ä¸ºå­—å…¸ (1 ä¸ªæŠ¥å‘Š)")
        
        print(f"æœ€ç»ˆç»“æœæ„å»ºå®Œæˆ: {len(all_results)} ä¸ªæ¼æ´æŠ¥å‘Š")
        print(f"  - æ‰€æœ‰æŠ¥å‘Šå‡ä¸ºç‹¬ç«‹æŠ¥å‘Šï¼ˆå·²å»é‡ï¼‰")
        
        return all_results

    def _clean_text_for_excel(self, text):
        """æ¸…ç†æ–‡æœ¬ä¸­çš„ç‰¹æ®Šå­—ç¬¦ï¼Œç¡®ä¿Excelå…¼å®¹æ€§"""
        if pd.isna(text):
            return ''
        
        # ç§»é™¤æˆ–æ›¿æ¢å¯èƒ½å¯¼è‡´Excelé—®é¢˜çš„å­—ç¬¦
        text = str(text).strip()
        # æ›¿æ¢å¸¸è§çš„ç‰¹æ®Šå­—ç¬¦
        replacements = {
            '\r': ' ',
            '\n': ' ',
            '\t': ' ',
            '\f': ' ',
            '\v': ' ',
            '\0': '',
            '\x00': '',
            '\u0000': '',
        }
        
        for old, new in replacements.items():
            text = text.replace(old, new)
        
        # ç§»é™¤å…¶ä»–ä¸å¯è§å­—ç¬¦
        text = ''.join(char for char in text if ord(char) >= 32 or char == ' ')
        
        return text

    def _extract_json_from_text(self, text):
        """ä»æ–‡æœ¬ä¸­æå–JSONå­—ç¬¦ä¸²ï¼Œå¢åŠ é˜²å¾¡æ€§ç¼–ç¨‹æªæ–½"""
        print("\nDebug - Starting JSON extraction")
        print(f"Debug - Input text length: {len(text)}")
        
        try:
            # é¦–å…ˆå°è¯•ç›´æ¥è§£ææ•´ä¸ªæ–‡æœ¬
            try:
                json.loads(text)
                return text
            except json.JSONDecodeError:
                pass
            
            # ç­–ç•¥1: æŸ¥æ‰¾æ ‡å‡†JSONä»£ç å—æ ‡è®°
            json_markers = ['```json', '`json', '```']
            for marker in json_markers:
                if marker in text:
                    # æ‰¾åˆ°æ ‡è®°åçš„å†…å®¹
                    start_pos = text.find(marker) + len(marker)
                    end_marker = '```' if marker.startswith('```') else '`'
                    end_pos = text.find(end_marker, start_pos)
                    
                    if end_pos != -1:
                        json_candidate = text[start_pos:end_pos].strip()
                        try:
                            json.loads(json_candidate)
                            print(f"Debug - Found JSON in {marker} block")
                            return json_candidate
                        except json.JSONDecodeError:
                            continue
            
            # ç­–ç•¥2: æŸ¥æ‰¾æ‰€æœ‰å¯èƒ½çš„JSONå¯¹è±¡
            json_patterns = [
                r'\{"[^"]+"\s*:\s*\[[^\]]*\][^}]*\}',  # æ ‡å‡†æ ¼å¼ {"key":["val1","val2"]}
                r'\{[^{}]*"group_[^"]*"[^{}]*\}',      # åŒ…å«group_çš„å¯¹è±¡
                r'\{[^{}]*:\s*\[[^\]]*\][^{}]*\}',     # ä»»ä½•key:arrayæ ¼å¼
            ]
            
            for pattern in json_patterns:
                matches = re.findall(pattern, text, re.DOTALL)
                for match in matches:
                    try:
                        # æ¸…ç†åŒ¹é…ç»“æœ
                        cleaned_match = match.strip()
                        json.loads(cleaned_match)
                        print(f"Debug - Found JSON with pattern: {pattern[:30]}...")
                        return cleaned_match
                    except json.JSONDecodeError:
                        continue
            
            # ç­–ç•¥3: æ‰‹åŠ¨æŸ¥æ‰¾æœ€åä¸€ä¸ªå®Œæ•´çš„JSONå¯¹è±¡
            # ä»æ–‡æœ¬æœ«å°¾å¼€å§‹æŸ¥æ‰¾}ï¼Œç„¶åå‘å‰æ‰¾åˆ°åŒ¹é…çš„{
            last_brace = text.rfind('}')
            if last_brace != -1:
                # ä»}ä½ç½®å‘å‰æŸ¥æ‰¾åŒ¹é…çš„{
                brace_count = 1
                start_pos = -1
                
                for i in range(last_brace - 1, -1, -1):
                    char = text[i]
                    if char == '}':
                        brace_count += 1
                    elif char == '{':
                        brace_count -= 1
                        if brace_count == 0:
                            start_pos = i
                            break
                
                if start_pos != -1:
                    json_candidate = text[start_pos:last_brace + 1]
                    try:
                        json.loads(json_candidate)
                        print(f"Debug - Found JSON by bracket matching")
                        return json_candidate
                    except json.JSONDecodeError:
                        pass
            
            # ç­–ç•¥4: æŸ¥æ‰¾åŒ…å«ç‰¹å®šå…³é”®è¯çš„è¡Œï¼Œå°è¯•æå–
            lines = text.split('\n')
            for line in lines:
                line = line.strip()
                if 'group_' in line and '{' in line and '}' in line:
                    try:
                        json.loads(line)
                        print(f"Debug - Found JSON in line with group_")
                        return line
                    except json.JSONDecodeError:
                        continue
            
            # ç­–ç•¥4.5: ä¸“é—¨å¤„ç†"æ­¥éª¤4"è¾“å‡ºæ ¼å¼
            # æŸ¥æ‰¾"æ­¥éª¤4"æˆ–ç±»ä¼¼æ ‡è®°åçš„JSON
            step_markers = ['æ­¥éª¤4', '**æ­¥éª¤4**', 'Step 4', 'è¾“å‡ºåˆ†ç±»ç»“æœ', 'åˆ†ç±»ç»“æœ']
            for marker in step_markers:
                if marker in text:
                    # æ‰¾åˆ°æ ‡è®°ä½ç½®
                    marker_pos = text.find(marker)
                    # ä»æ ‡è®°åå¼€å§‹æŸ¥æ‰¾JSON
                    remaining_text = text[marker_pos:]
                    
                    # åœ¨å‰©ä½™æ–‡æœ¬ä¸­æŸ¥æ‰¾JSON
                    if '{' in remaining_text and '}' in remaining_text:
                        start = remaining_text.find('{')
                        end = remaining_text.rfind('}') + 1
                        json_candidate = remaining_text[start:end]
                        
                        try:
                            json.loads(json_candidate)
                            print(f"Debug - Found JSON after {marker}")
                            return json_candidate
                        except json.JSONDecodeError:
                            # å°è¯•æå–æœ€åä¸€è¡Œçš„JSON
                            lines_after_marker = remaining_text[start:].split('\n')
                            for line in lines_after_marker:
                                line = line.strip()
                                if line.startswith('{') and line.endswith('}'):
                                    try:
                                        json.loads(line)
                                        print(f"Debug - Found JSON in line after {marker}")
                                        return line
                                    except json.JSONDecodeError:
                                        continue
            
            # ç­–ç•¥5: å°è¯•ä¿®å¤å¸¸è§çš„JSONæ ¼å¼é—®é¢˜
            # æŸ¥æ‰¾çœ‹èµ·æ¥åƒJSONçš„å†…å®¹å¹¶å°è¯•ä¿®å¤
            potential_json = None
            if '{' in text and '}' in text:
                start = text.find('{')
                end = text.rfind('}') + 1
                potential_json = text[start:end]
                
                # å°è¯•ä¸€äº›å¸¸è§çš„ä¿®å¤
                fixes = [
                    lambda x: x,  # åŸæ ·
                    lambda x: x.replace("'", '"'),  # å•å¼•å·æ”¹åŒå¼•å·
                    lambda x: re.sub(r'(\w+):', r'"\1":', x),  # ç»™keyåŠ å¼•å·
                    lambda x: re.sub(r':\s*([^",\[\]{}]+)(?=[,}])', r': "\1"', x),  # ç»™valueåŠ å¼•å·
                ]
                
                for fix_func in fixes:
                    try:
                        fixed_json = fix_func(potential_json)
                        json.loads(fixed_json)
                        print(f"Debug - Fixed JSON format")
                        return fixed_json
                    except (json.JSONDecodeError, Exception):
                        continue
            
            raise ValueError("No valid JSON object found after all strategies")
            
        except (json.JSONDecodeError, ValueError) as e:
            print(f"Debug - Error extracting JSON: {str(e)}")
            print(f"Debug - Original text: {text}")
            
            # æœ€åçš„å¤‡ç”¨ç­–ç•¥ï¼šå°è¯•æ„é€ ä¸€ä¸ªç®€å•çš„åˆ†ç»„
            print("Debug - Attempting fallback JSON construction")
            try:
                # å¦‚æœæ–‡æœ¬ä¸­æåˆ°äº†åˆ†ç»„ï¼Œå°è¯•æ„é€ ä¸€ä¸ªç®€å•çš„JSON
                if 'group_' in text.lower():
                    fallback_json = '{"group_1":[]}'
                    print("Debug - Using fallback JSON")
                    return fallback_json
            except:
                pass
            
            raise ValueError(f"Failed to extract valid JSON after all strategies: {str(e)}")

    def _translate_to_chinese(self, final_results):
        """å°†æ¼æ´ç»“æœç¿»è¯‘æˆä¸­æ–‡"""
        if not self.enable_chinese_translation:
            print("ä¸­æ–‡ç¿»è¯‘åŠŸèƒ½æœªå¯ç”¨ï¼Œè·³è¿‡ç¿»è¯‘")
            return final_results
        
        print(f"\n=== å¼€å§‹ä¸­æ–‡ç¿»è¯‘ ===")
        print(f"éœ€è¦ç¿»è¯‘ {len(final_results)} ä¸ªæ¼æ´ç»“æœ...")
        
        translated_results = []
        
        with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
            future_to_index = {
                executor.submit(self._translate_single_result, i, result): i 
                for i, result in enumerate(final_results)
            }
            
            with tqdm(total=len(final_results), desc="ä¸­æ–‡ç¿»è¯‘") as pbar:
                for future in concurrent.futures.as_completed(future_to_index):
                    try:
                        index = future_to_index[future]
                        translated_result = future.result()
                        
                        # ä¿æŒåŸæœ‰é¡ºåº
                        while len(translated_results) <= index:
                            translated_results.append(None)
                        translated_results[index] = translated_result
                        
                        print(f"  æ¼æ´ {index+1} ç¿»è¯‘å®Œæˆ")
                    except Exception as e:
                        index = future_to_index[future]
                        original_result = final_results[index]
                        print(f"  æ¼æ´ {index+1} ç¿»è¯‘å¤±è´¥: {str(e)}")
                        
                        # ä¿æŒåŸæœ‰é¡ºåº
                        while len(translated_results) <= index:
                            translated_results.append(None)
                        translated_results[index] = original_result
                    pbar.update(1)
        
        # è¿‡æ»¤æ‰Noneå€¼å¹¶ä¿æŒé¡ºåº
        translated_results = [result for result in translated_results if result is not None]
        
        print(f"ä¸­æ–‡ç¿»è¯‘å®Œæˆ: {len(translated_results)} ä¸ªæ¼æ´ç»“æœ")
        return translated_results

    def _translate_single_result(self, index, result):
        """ç¿»è¯‘å•ä¸ªæ¼æ´ç»“æœ"""
        original_result = result.get('æ¼æ´ç»“æœ', '')
        
        if not original_result or pd.isna(original_result):
            print(f"    æ¼æ´ {index+1}: æ— å†…å®¹ï¼Œè·³è¿‡ç¿»è¯‘")
            return result
        
        translate_prompt = f"""å°†è¿™ä¸ªæ¼æ´è¯¦ç»†çš„ç”¨ä¸­æ–‡è§£é‡Šä¸€ä¸‹ï¼Œä¸è¦é—æ¼ä»»ä½•ç»†èŠ‚

æ¼æ´æè¿°ï¼š
{original_result}"""

        max_retries = 3  # æœ€å¤§é‡è¯•æ¬¡æ•°
        retry_count = 0
        
        while retry_count < max_retries:
            try:
                if retry_count == 0:
                    print(f"    æ¼æ´ {index+1}: å¼€å§‹ç¿»è¯‘...")
                else:
                    print(f"    æ¼æ´ {index+1}: ç¬¬ {retry_count+1} æ¬¡é‡è¯•ç¿»è¯‘...")
                
                translated_description = ask_claude(translate_prompt)
                
                # æ¸…ç†ç¿»è¯‘ç»“æœ
                cleaned_description = self._clean_text_for_excel(translated_description)
                
                # æ£€æŸ¥ç¿»è¯‘åçš„é•¿åº¦
                if len(cleaned_description) == 0:
                    print(f"    æ¼æ´ {index+1}: âš ï¸  ç¿»è¯‘åé•¿åº¦ä¸º0ï¼ŒåŸé•¿åº¦ {len(original_result)}")
                    retry_count += 1
                    if retry_count < max_retries:
                        print(f"    æ¼æ´ {index+1}: å‡†å¤‡é‡æ–°ç¿»è¯‘ (ç¬¬ {retry_count+1} æ¬¡å°è¯•)")
                        continue
                    else:
                        print(f"    æ¼æ´ {index+1}: âŒ é‡è¯• {max_retries} æ¬¡åä»ç„¶å¤±è´¥ï¼Œä¿ç•™åŸç»“æœ")
                        return result
                
                # åˆ›å»ºæ–°çš„ç»“æœå‰¯æœ¬
                translated_result = result.copy()
                translated_result['æ¼æ´ç»“æœ'] = cleaned_description
                
                print(f"    æ¼æ´ {index+1}: ç¿»è¯‘æˆåŠŸï¼ŒåŸé•¿åº¦ {len(original_result)} -> æ–°é•¿åº¦ {len(cleaned_description)}")
                
                return translated_result
                
            except Exception as e:
                print(f"    æ¼æ´ {index+1}: ç¿»è¯‘å¤±è´¥ - {str(e)}")
                retry_count += 1
                if retry_count < max_retries:
                    print(f"    æ¼æ´ {index+1}: å‡†å¤‡é‡è¯• (ç¬¬ {retry_count+1} æ¬¡å°è¯•)")
                    continue
                else:
                    print(f"    æ¼æ´ {index+1}: âŒ é‡è¯• {max_retries} æ¬¡åä»ç„¶å¤±è´¥ï¼Œä¿ç•™åŸç»“æœ")
                    return result
        
        # å¦‚æœæ‰€æœ‰é‡è¯•éƒ½å¤±è´¥ï¼Œè¿”å›åŸç»“æœ
        return result

    @staticmethod
    def perform_post_reasoning_deduplication(project_id, db_engine, logger):
        """åœ¨reasoningå®Œæˆåï¼Œvalidationå¼€å§‹å‰è¿›è¡Œå»é‡å¤„ç†"""
        from logging_config import log_step, log_section_start, log_section_end, log_error, log_warning, log_success, log_data_info
        from dao import ProjectTaskMgr
        
        log_step(logger, "å¼€å§‹è·å–reasoningåçš„æ¼æ´æ•°æ®")
        
        try:
            # è·å–reasoningåçš„æ‰€æœ‰æ¼æ´æ•°æ®
            project_taskmgr = ProjectTaskMgr(project_id, db_engine)
            entities = project_taskmgr.query_task_by_project_id(project_id)
            
            # è°ƒè¯•ä¿¡æ¯ï¼šç»Ÿè®¡æ‰€æœ‰å®ä½“
            total_entities = len(entities)
            log_data_info(logger, "æ€»ä»»åŠ¡å®ä½“æ•°é‡", total_entities)
            print(f"ğŸ” è°ƒè¯•ä¿¡æ¯ - æ€»ä»»åŠ¡å®ä½“æ•°é‡: {total_entities}")
            
            # ç­›é€‰æœ‰æ¼æ´ç»“æœçš„æ•°æ®
            vulnerability_data = []
            for entity in entities:
                # è°ƒè¯•æ¯ä¸ªå®ä½“çš„è¯¦ç»†ä¿¡æ¯
                has_result = bool(entity.result)
                has_business_code = hasattr(entity, 'business_flow_code') and entity.business_flow_code and len(entity.business_flow_code) > 0
                
                if has_result and has_business_code:
                    vulnerability_data.append({
                        'æ¼æ´ç»“æœ': entity.result,
                        'ID': entity.id,
                        'é¡¹ç›®åç§°': entity.project_id,
                        'åˆåŒç¼–å·': entity.contract_code,
                        'UUID': entity.uuid,
                        'å‡½æ•°åç§°': entity.name,
                        'å‡½æ•°ä»£ç ': entity.content,
                        'è§„åˆ™ç±»å‹': entity.rule_key,
                        'å¼€å§‹è¡Œ': entity.start_line,
                        'ç»“æŸè¡Œ': entity.end_line,
                        'ç›¸å¯¹è·¯å¾„': entity.relative_file_path,
                        'ç»å¯¹è·¯å¾„': entity.absolute_file_path,
                        'ä¸šåŠ¡æµç¨‹ä»£ç ': entity.business_flow_code,
                        'æ‰«æè®°å½•': entity.scan_record,
                        'æ¨è': entity.recommendation
                    })
            
            filtered_count = len(vulnerability_data)
            print(f"ğŸ” è°ƒè¯•ä¿¡æ¯ - é€šè¿‡ç­›é€‰æ¡ä»¶çš„å®ä½“: {filtered_count}")
                        
            original_df = pd.DataFrame(vulnerability_data)
            original_count = len(original_df)
            original_ids = set(original_df['ID'].astype(str))
            
            log_data_info(logger, "å»é‡å‰æ¼æ´æ•°é‡", original_count)
            log_data_info(logger, "å»é‡å‰æ¼æ´ID", f"{', '.join(sorted(original_ids))}")
            
            # æ£€æŸ¥æ˜¯å¦å­˜åœ¨å·²ç»è¢«é€»è¾‘åˆ é™¤çš„è®°å½•ï¼ˆshort_result='delete'ï¼‰
            all_entities = project_taskmgr.query_task_by_project_id(project_id)
            deleted_tasks = [entity for entity in all_entities if getattr(entity, 'short_result', '') == 'delete']
            
            if deleted_tasks:
                deleted_count = len(deleted_tasks)
                deleted_ids = [str(task.id) for task in deleted_tasks]
                print(f"\nâš ï¸  æ£€æµ‹åˆ°å·²æœ‰ {deleted_count} ä¸ªé€»è¾‘åˆ é™¤çš„è®°å½•ï¼Œè·³è¿‡ResProcessorå»é‡å¤„ç†")
                print(f"å·²åˆ é™¤çš„ID: {', '.join(deleted_ids)}")
                log_warning(logger, f"è·³è¿‡ResProcessorå»é‡å¤„ç† - æ£€æµ‹åˆ°{deleted_count}ä¸ªå·²é€»è¾‘åˆ é™¤çš„è®°å½•")
                log_warning(logger, f"å·²åˆ é™¤çš„ID: {', '.join(deleted_ids)}")
                return
            
            # ä½¿ç”¨ResProcessorè¿›è¡Œå»é‡
            log_step(logger, "å¼€å§‹ResProcessorå»é‡å¤„ç†")
            res_processor = ResProcessor(original_df, max_group_size=5, iteration_rounds=8, enable_chinese_translation=False)
            processed_df = res_processor.process()
            
            deduplicated_count = len(processed_df)
            deduplicated_ids = set(processed_df['ID'].astype(str))
            
            log_data_info(logger, "å»é‡åæ¼æ´æ•°é‡", deduplicated_count)
            log_data_info(logger, "å»é‡åæ¼æ´ID", f"{', '.join(sorted(deduplicated_ids))}")
            
            # è®¡ç®—è¢«å»é‡çš„ID
            removed_ids = original_ids - deduplicated_ids
            removed_count = len(removed_ids)
            
            # æ‰“å°å»é‡ç»“æœ
            print(f"\n{'='*60}")
            print(f"ğŸ”„ Reasoningåå»é‡å¤„ç†ç»“æœ")
            print(f"{'='*60}")
            print(f"å»é‡å‰æ¼æ´æ•°é‡: {original_count}")
            print(f"å»é‡åæ¼æ´æ•°é‡: {deduplicated_count}")
            print(f"è¢«å»é‡çš„æ¼æ´æ•°é‡: {removed_count}")
            
            if removed_ids:
                print(f"\nğŸ—‘ï¸  è¢«å»é‡çš„æ¼æ´IDåˆ—è¡¨:")
                for i, removed_id in enumerate(sorted(removed_ids), 1):
                    print(f"  {i:2d}. ID: {removed_id}")
                
                # é€»è¾‘åˆ é™¤è¢«å»é‡çš„è®°å½• - å°†short_resultè®¾ç½®ä¸º"delete"
                print(f"\nğŸ—‘ï¸  å¼€å§‹é€»è¾‘åˆ é™¤è¢«å»é‡çš„è®°å½•(è®¾ç½®short_result='delete')...")
                marked_count = 0
                failed_marks = []
                
                for removed_id in removed_ids:
                    try:
                        # è½¬æ¢ä¸ºæ•´æ•°ç±»å‹çš„ID
                        id_int = int(removed_id)
                        project_taskmgr.update_short_result(id_int, "delete")
                        marked_count += 1
                        print(f"    âœ… æ ‡è®°æˆåŠŸ: ID {removed_id} -> short_result='delete'")
                    except Exception as e:
                        failed_marks.append(removed_id)
                        print(f"    âŒ æ ‡è®°å‡ºé”™: ID {removed_id}, é”™è¯¯: {str(e)}")
                        logger.error(f"æ ‡è®°åˆ é™¤ID {removed_id} æ—¶å‡ºé”™: {str(e)}")
                
                print(f"\nğŸ“Š é€»è¾‘åˆ é™¤ç»“æœ:")
                print(f"    æˆåŠŸæ ‡è®°: {marked_count} æ¡è®°å½•")
                if failed_marks:
                    print(f"    æ ‡è®°å¤±è´¥: {len(failed_marks)} æ¡è®°å½• - IDs: {', '.join(failed_marks)}")
                    logger.warning(f"æ ‡è®°å¤±è´¥çš„IDs: {', '.join(failed_marks)}")
                
                log_success(logger, "é€»è¾‘åˆ é™¤æ“ä½œå®Œæˆ", f"æˆåŠŸæ ‡è®°: {marked_count}/{removed_count}")
            else:
                print("âœ… æ²¡æœ‰æ¼æ´è¢«å»é‡")
            
            print(f"{'='*60}\n")
            
            # è®°å½•åˆ°æ—¥å¿—
            log_success(logger, "å»é‡å¤„ç†å®Œæˆ", f"åŸå§‹: {original_count} -> å»é‡å: {deduplicated_count}, é€»è¾‘åˆ é™¤: {removed_count}")
            if removed_ids:
                logger.info(f"è¢«å»é‡çš„æ¼æ´ID: {', '.join(sorted(removed_ids))}")
                logger.info(f"é€»è¾‘åˆ é™¤äº† {marked_count} æ¡è¢«å»é‡çš„è®°å½•(è®¾ç½®short_result='delete')")
            
        except Exception as e:
            log_error(logger, "å»é‡å¤„ç†å¤±è´¥", e)
            import traceback
            logger.error(f"è¯¦ç»†é”™è¯¯ä¿¡æ¯: {traceback.format_exc()}")

    @staticmethod
    def generate_excel(output_path, project_id, db_engine):
        """ç”ŸæˆExcelæŠ¥å‘Š"""
        from dao import ProjectTaskMgr
        
        project_taskmgr = ProjectTaskMgr(project_id, db_engine)
        entities = project_taskmgr.query_task_by_project_id(project_id)
        
        # åˆ›å»ºä¸€ä¸ªç©ºçš„DataFrameæ¥å­˜å‚¨æ‰€æœ‰å®ä½“çš„æ•°æ®
        data = []
        total_entities = len(entities)
        deleted_entities = 0
        
        for entity in entities:
            # è·³è¿‡å·²é€»è¾‘åˆ é™¤çš„è®°å½•
            if getattr(entity, 'short_result', '') == 'delete':
                deleted_entities += 1
                continue
                
            # ä¼˜å…ˆä½¿ç”¨validationåçš„short_resultï¼Œå¦‚æœæ²¡æœ‰åˆ™ä½¿ç”¨åŸå§‹result
            short_result = entity.short_result
            result = entity.result
            if short_result and ("yes" in str(short_result).lower()) and len(entity.business_flow_code)>0:
                data.append({
                    'æ¼æ´ç»“æœ': result,
                    'ID': entity.id,
                    'é¡¹ç›®åç§°': entity.project_id,
                    'åˆåŒç¼–å·': entity.contract_code,
                    'UUID': entity.uuid,  # ä½¿ç”¨uuidè€Œä¸æ˜¯key
                    'å‡½æ•°åç§°': entity.name,
                    'å‡½æ•°ä»£ç ': entity.content,
                    'è§„åˆ™ç±»å‹': entity.rule_key,  # æ–°å¢rule_key
                    'å¼€å§‹è¡Œ': entity.start_line,
                    'ç»“æŸè¡Œ': entity.end_line,
                    'ç›¸å¯¹è·¯å¾„': entity.relative_file_path,
                    'ç»å¯¹è·¯å¾„': entity.absolute_file_path,
                    'ä¸šåŠ¡æµç¨‹ä»£ç ': entity.business_flow_code,
                    'æ‰«æè®°å½•': entity.scan_record,  # ä½¿ç”¨æ–°çš„scan_recordå­—æ®µ
                    'æ¨è': entity.recommendation
                })
        
        # æ‰“å°æ•°æ®ç»Ÿè®¡ä¿¡æ¯
        print(f"\nğŸ“Š ExcelæŠ¥å‘Šæ•°æ®ç»Ÿè®¡:")
        print(f"   æ€»è®°å½•æ•°: {total_entities}")
        print(f"   é€»è¾‘åˆ é™¤çš„è®°å½•æ•°: {deleted_entities}")
        print(f"   æœ‰æ•ˆè®°å½•æ•°: {total_entities - deleted_entities}")
        print(f"   ç¬¦åˆæ¡ä»¶çš„æ¼æ´è®°å½•æ•°: {len(data)}")
        
        # å°†æ•°æ®è½¬æ¢ä¸ºDataFrame
        if not data:  # æ£€æŸ¥æ˜¯å¦æœ‰æ•°æ®
            print("No data to process")
            return
            
        df = pd.DataFrame(data)
        
        try:
            # å¯¹dfè¿›è¡Œæ¼æ´å½’é›†å¤„ç†
            res_processor = ResProcessor(df, max_group_size=10, iteration_rounds=8, enable_chinese_translation=True)
            processed_df = res_processor.process()
            
            # ç¡®ä¿æ‰€æœ‰å¿…éœ€çš„åˆ—éƒ½å­˜åœ¨
            required_columns = df.columns
            for col in required_columns:
                if col not in processed_df.columns:
                    processed_df[col] = ''
                    
            # é‡æ–°æ’åˆ—åˆ—é¡ºåºä»¥åŒ¹é…åŸå§‹DataFrame
            processed_df = processed_df[df.columns]
        except Exception as e:
            print(f"Error processing data: {e}")
            processed_df = df  # å¦‚æœå¤„ç†å¤±è´¥ï¼Œä½¿ç”¨åŸå§‹DataFrame
        
        # ç¡®ä¿è¾“å‡ºç›®å½•å­˜åœ¨
        output_dir = os.path.dirname(output_path)
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        # æ£€æŸ¥æ–‡ä»¶æ˜¯å¦å­˜åœ¨ï¼Œå¦‚æœä¸å­˜åœ¨åˆ™åˆ›å»ºæ–°æ–‡ä»¶
        if not os.path.exists(output_path):
            wb = Workbook()
            ws = wb.active
            ws.title = "é¡¹ç›®æ•°æ®"
        else:
            wb = load_workbook(output_path)
            if "é¡¹ç›®æ•°æ®" in wb.sheetnames:
                ws = wb["é¡¹ç›®æ•°æ®"]
            else:
                ws = wb.create_sheet("é¡¹ç›®æ•°æ®")
        
        # å¦‚æœå·¥ä½œè¡¨æ˜¯ç©ºçš„ï¼Œæ·»åŠ è¡¨å¤´
        if ws.max_row == 1:
            for col, header in enumerate(processed_df.columns, start=1):
                ws.cell(row=1, column=col, value=header)
        
        # å°†DataFrameæ•°æ®å†™å…¥å·¥ä½œè¡¨
        for row in dataframe_to_rows(processed_df, index=False, header=False):
            ws.append(row)
        
        # ä¿å­˜Excelæ–‡ä»¶
        wb.save(output_path)
        
        print(f"Excelæ–‡ä»¶å·²ä¿å­˜åˆ°: {output_path}")